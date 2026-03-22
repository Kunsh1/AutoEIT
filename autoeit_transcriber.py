"""
AutoEIT Transcriber
────────────────────
Reads speaker_separator.py output (transcript.txt per file) and maps
each participant response to one of the 30 known EIT sentences, then
writes the result into column C of the Excel file.

Usage:
    python autoeit_transcriber.py
    python autoeit_transcriber.py --transcripts ./speaker_output --excel ./AutoEIT_Sample_Audio_for_Transcribing.xlsx

Requirements:
    pip install rapidfuzz openpyxl langdetect
"""

import os
import sys
import re
import argparse

try:
    from langdetect import detect as _langdetect
    _LANGDETECT_AVAILABLE = True
except ImportError:
    _LANGDETECT_AVAILABLE = False

# ─── CONFIG ───────────────────────────────────────────────────────────────────

TRANSCRIPTS_ROOT = "./speaker_output"
EXCEL_PATH       = "./AutoEIT Sample Audio for Transcribing.xlsx"

# Threshold used to decide if a segment is a completely new attempt (OVERWRITE)
# or just a continuation of the previous transcript chunk (CONCATENATE).
PAIR_THRESHOLD = 55

# ─── THE 30 KNOWN EIT STIMULI ─────────────────────────────────────────────────

EIT_STIMULI_ENGLISH = [
    "We drove to the park",
    "I'll call her tomorrow night",
    "You can buy meat at the butcher shop",
    "My brother just bought a brand new computer",
    "Sometimes they take their dog for a walk in the park",
    "We're going to play volleyball at the gym that I told you about",
]

EIT_STIMULI_SPANISH = [
    "Quiero cortarme el pelo",
    "El libro está en la mesa",
    "El carro lo tiene Pedro",
    "Él se ducha cada mañana",
    "¿Qué dice usted que va a hacer hoy?",
    "Dudo que sepa manejar muy bien",
    "Las calles de esta ciudad son muy anchas",
    "Puede que llueva mañana todo el día",
    "Las casas son muy bonitas pero caras",
    "Me gustan las películas que acaban bien",
    "El chico con el que yo salgo es español",
    "Después de cenar me fui a dormir tranquilo",
    "Quiero una casa en la que vivan mis animales",
    "A nosotros nos fascinan las fiestas grandiosas",
    "Ella sólo bebe cerveza y no come nada",
    "Me gustaría que el precio de las casas bajara",
    "Cruza a la derecha y después sigue todo recto",
    "Ella ha terminado de pintar su apartamento",
    "Me gustaría que empezara a hacer más calor pronto",
    "El niño al que se le murió el gato está triste",
    "Una amiga mía cuida a los niños de mi vecino",
    "El gato que era negro fue perseguido por el perro",
    "Antes de poder salir él tiene que limpiar su cuarto",
    "La cantidad de personas que fuman ha disminuido",
    "Después de llegar a casa del trabajo tomé la cena",
    "El ladrón al que atrapó la policía era famoso",
    "Le pedí a un amigo que me ayudara con la tarea",
    "El examen no fue tan difícil como me habían dicho",
    "¿Serías tan amable de darme el libro que está en la mesa?",
    "Hay mucha gente que no toma nada para el desayuno",
]
# ──────────────────────────────────────────────────────────────────────────────

COLORS = {
    "reset":  "\033[0m",  "bold":   "\033[1m",  "dim":    "\033[2m",
    "header": "\033[96m", "green":  "\033[92m",  "red":    "\033[91m",
    "yellow": "\033[93m", "blue":   "\033[94m",  "cyan":   "\033[96m",
}

def color(text, *keys):
    codes = "".join(COLORS.get(k, "") for k in keys)
    return f"{codes}{text}{COLORS['reset']}"

def print_header(title):
    print("\n" + color("─" * 70, "header"))
    print(color(f"  {title}", "header", "bold"))
    print(color("─" * 70, "header"))

def print_banner(title):
    print("\n" + color("═" * 70, "bold"))
    print(color(f"  {title}", "bold"))
    print(color("═" * 70, "bold"))


def check_requirements():
    missing = []
    for pkg, imp in [("rapidfuzz", "rapidfuzz"), ("openpyxl", "openpyxl")]:
        try:
            __import__(imp)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(color(f"\n❌ Missing packages. Run:\n\n   pip install {' '.join(missing)}\n", "bold"))
        sys.exit(1)


import unicodedata

def normalize(text):
    text = text.lower()
    # Strip accents
    text = ''.join(c for c in unicodedata.normalize('NFD', text)
                   if unicodedata.category(c) != 'Mn')
    text = re.sub(r"[¿¡.,!?;:\"'«»\-]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def fuzzy_score(a, b):
    from rapidfuzz import fuzz
    return fuzz.token_sort_ratio(normalize(a), normalize(b))


def best_stimulus_match(text, stimuli):
    """Return (best_stimulus, best_index_0based, best_score)."""
    best_score, best_idx, best_stim = -1, -1, ""
    for i, stim in enumerate(stimuli):
        score = fuzzy_score(text, stim)
        if score > best_score:
            best_score, best_idx, best_stim = score, i, stim
    return best_stim, best_idx, best_score


def is_english(text):
    """
    Return True if the segment is clearly English.
    Safety net to discard instruction/practice sentences
    even if boundary detection fails.
    Ignores segments < 2 words to avoid false positives.
    """
    if len(text.split()) < 2:
        return False
    if _LANGDETECT_AVAILABLE:
        try:
            return _langdetect(text) == "en"
        except Exception:
            return False
    # Fallback keyword check if langdetect not installed
    english_markers = {"the", "you", "your", "will", "please", "dont",
                       "after", "each", "until", "start", "hear", "tone",
                       "sentence", "repeat", "given", "sufficient", "careful",
                       "notes", "recording", "attention", "exercise", "bought",
                       "drive", "drove", "park", "call", "buy", "meat", "butcher",
                       "brother", "computer", "sometimes", "volleyball", "gym"}
    words = set(normalize(text).split())
    return len(words & english_markers) >= 2


# ─── SKIP PHRASES ─────────────────────────────────────────────────────────────
# Two categories:
#   1. EIT protocol instructions (Spanish + English)
#   2. Whisper prompt leak phrases — Whisper sometimes hallucinates the
#      initial_prompt text verbatim into a segment. Any segment containing
#      these substrings is ASR noise, not a participant response.

SPANISH_INSTRUCTION_PHRASES = [
    # ── EIT protocol instructions ──────────────────────────────────────────
    "repite lo máximo que puedas",
    "repite lo más que puedas",
    "recuerda no comiences",
    "no comiences a repetir",
    "te darás suficiente tiempo",
    "darás suficiente tiempo",
    "después del tono para repetir",
    "escuches el tono",
    "hasta que escuches",
    "ahora volvemos",
    "ahora comenzamos",
    "vamos a comenzar",
    "ahora empezamos",
    "presta atención",
    "por favor no tomes",
    "no tomes notas",
    "no anotes nada",
    "now let s begin",
    "now let us begin",
    "let s begin",
    # ── Whisper prompt leak phrases ────────────────────────────────────────
    # These are substrings from the INITIAL_PROMPT in speaker_separator.py.
    # If any appear in a segment it means Whisper hallucinated the prompt,
    # not a real participant response.
    "incluye errores gramaticales",
    "transcribe exactamente lo que dice",
    "falsos inicios con guion",
    "falsos inicios con guión",
    "no corrijas ni normalices",
    "no traduzcas ninguna frase",
    "si hay una pausa demasiado larga",
    "probablemente indica el inicio de la siguiente",
    "un estudiante de español l2 repite",
    "estudiante de español l2 repite una frase",
]


def should_skip(text):
    """
    Return True if a segment should be discarded entirely:
      - Clearly English (instruction / English practice section)
      - Matches a known EIT instruction phrase
      - Contains a Whisper prompt leak phrase
    """
    if is_english(text):
        return True
    text_norm = normalize(text)
    for phrase in SPANISH_INSTRUCTION_PHRASES:
        if phrase in text_norm:
            return True
    return Falsedef parse_transcript(transcript_path):
    """
    Parse transcript.txt → list of {start_sec, end_sec, text}.
    Handles lines like:  [02:34 → 02:37]  Speaker 1   El libro está...
    """
    segments = []
    pattern  = re.compile(r"\[(\d{2}:\d{2})\s*[→\-]+\s*(\d{2}:\d{2})\]\s*(.*)")

    def to_sec(t):
        mm, ss = t.split(":")
        return int(mm) * 60 + int(ss)

    with open(transcript_path, encoding="utf-8") as f:
        for line in f:
            m = pattern.search(line.strip())
            if not m:
                continue
            raw_text = m.group(3).strip()
            text     = re.sub(r"^(Speaker\s*\d+|Unknown)\s+", "", raw_text, flags=re.IGNORECASE).strip()

            # ── NEW: Fix Whisper's missing spaces after punctuation ──
            # Turns "tranquilo.Quiero" into "tranquilo. Quiero"
            text = re.sub(r'([.?!])([A-Za-z¿¡])', r'\1 \2', text)
            # ─────────────────────────────────────────────────────────

            if text:
                segments.append({
                    "start": to_sec(m.group(1)),
                    "end":   to_sec(m.group(2)),
                    "text":  text,
                })
    return segments


def find_spanish_boundary(segments):
    """
    Find the first segment index where the Spanish EIT trials start.
    """
    if len(segments) < 2:
        return 0

    total_duration = segments[-1]["end"]
    search_limit   = total_duration * 0.35

    # ── Priority 1: last English segment in first 35% ─────────────────────
    last_english_idx = -1
    for i, seg in enumerate(segments):
        if seg["start"] > search_limit:
            break
        if is_english(seg["text"]) or should_skip(seg["text"]):
            last_english_idx = i

    if last_english_idx >= 0:
        boundary = last_english_idx + 1
        preview  = segments[last_english_idx]["text"][:50]
        ts       = segments[last_english_idx]["start"]
        print(f"    ✅ English/Spanish boundary: after segment #{last_english_idx} "
              f"(last English at {ts}s: '{preview}')")
        return boundary

    # ── Priority 2: largest silence gap in first 35% ──────────────────────
    best_gap, best_idx = 0, 0
    for i in range(1, len(segments)):
        start_prev = segments[i-1]["start"]
        if start_prev < 60:
            continue
        if start_prev > search_limit:
            break
        gap = segments[i]["start"] - segments[i-1]["end"]
        if gap > best_gap:
            best_gap, best_idx = gap, i

    if best_idx > 0:
        print(f"    ✅ English/Spanish boundary: after segment #{best_idx} "
              f"(gap={best_gap:.0f}s  at {segments[best_idx]['start']}s)")
        return best_idx

    # ── Priority 3: fallback — absolute last English segment ──────────────
    last_absolute_english = -1
    for i, seg in enumerate(segments):
        if is_english(seg["text"]) or should_skip(seg["text"]):
            last_absolute_english = i

    if last_absolute_english >= 0:
        print(f"    ⚠️  Boundary fallback: using absolute last English segment (#{last_absolute_english})")
        return last_absolute_english + 1

    return 0def split_merged_segment(text, stimuli):
    """
    Detect and split segments containing MULTIPLE merged responses.
    Uses recursion to handle 3+ sentences squeezed into one segment.
    """
    if "..." in text:
        parts_on_ellipsis = text.split("...")
        candidates = []
        for cut in range(1, len(parts_on_ellipsis)):
            left  = "...".join(parts_on_ellipsis[:cut]).strip().rstrip(".")
            right = "...".join(parts_on_ellipsis[cut:]).strip().lstrip(".")
            if len(left.split()) >= 2 and len(right.split()) >= 2:
                candidates.append((left, right))

        best_score, best_pair = 0, None
        _, _, single_score = best_stimulus_match(text, stimuli)

        for left, right in candidates:
            _, li, ls = best_stimulus_match(left,  stimuli)
            _, ri, rs = best_stimulus_match(right, stimuli)

            # THE FIX: Ensure the right side doesn't time-travel backward
            if li != ri and ls >= 40 and rs >= 40 and ri >= li - 2:
                combined = ls + rs
                if combined > best_score:
                    best_score, best_pair = combined, (left, right)

        if best_pair and best_score > single_score + 15:
            return split_merged_segment(best_pair[0], stimuli) + split_merged_segment(best_pair[1], stimuli)

    words = text.split()
    if len(words) < 8:
        return [text]

    _, _, single_score = best_stimulus_match(text, stimuli)
    best_combined, best_pair = 0, None

    for split in range(3, len(words) - 2):
        left  = " ".join(words[:split])
        right = " ".join(words[split:])
        _, li, ls = best_stimulus_match(left,  stimuli)
        _, ri, rs = best_stimulus_match(right, stimuli)

        # THE FIX: Ensure the right side doesn't time-travel backward wildly (ri >= li - 2)
        if li != ri and ls >= 60 and rs >= 60 and ri >= li - 2:
            combined = ls + rs
            if combined > best_combined:
                best_combined, best_pair = combined, (left, right)

    if best_pair and best_combined > single_score + 30:
        return split_merged_segment(best_pair[0], stimuli) + split_merged_segment(best_pair[1], stimuli)

    return [text]

def fix_bleed_over(responses, stimuli):
    """
    Shifts trailing words that belong to the next sentence (Forward bleed)
    AND leading words that belong to the previous sentence (Backward bleed).
    """
    # ── 1. FORWARD BLEED (tail of current belongs to head of next) ──
    for i in range(1, 30):
        if i in responses and (i + 1) in responses:
            curr_words = responses[i].split()
            next_stim_words = normalize(stimuli[i]).split() # stimuli is 0-indexed

            if not curr_words or not next_stim_words:
                continue

            # Check 2-word forward bleed
            if len(curr_words) >= 2 and len(next_stim_words) >= 2:
                tail_2 = normalize(curr_words[-2] + " " + curr_words[-1])
                head_2 = next_stim_words[0] + " " + next_stim_words[1]
                if tail_2 == head_2:
                    w2 = curr_words.pop()
                    w1 = curr_words.pop()
                    responses[i]   = " ".join(curr_words)
                    responses[i+1] = f"{w1} {w2} {responses[i+1]}"
                    continue

            # Check 1-word forward bleed
            if len(curr_words) >= 1 and len(next_stim_words) >= 1:
                tail_1 = normalize(curr_words[-1])
                head_1 = next_stim_words[0]
                if tail_1 == head_1:
                    w1 = curr_words.pop()
                    responses[i]   = " ".join(curr_words)
                    responses[i+1] = f"{w1} {responses[i+1]}"

    # ── 2. BACKWARD BLEED (head of next belongs to tail of current) ──
    for i in range(1, 30):
        if i in responses and (i + 1) in responses:
            next_words = responses[i+1].split()
            curr_stim_words = normalize(stimuli[i-1]).split()

            if not next_words or not curr_stim_words:
                continue

            # Check 2-word backward bleed
            if len(next_words) >= 2 and len(curr_stim_words) >= 2:
                head_2 = normalize(next_words[0] + " " + next_words[1])
                tail_2 = curr_stim_words[-2] + " " + curr_stim_words[-1]
                if head_2 == tail_2:
                    w1 = next_words.pop(0)
                    w2 = next_words.pop(0)
                    responses[i]   = f"{responses[i]} {w1} {w2}"
                    responses[i+1] = " ".join(next_words)
                    continue

            # Check 1-word backward bleed
            if len(next_words) >= 1 and len(curr_stim_words) >= 1:
                head_1 = normalize(next_words[0])
                tail_1 = curr_stim_words[-1]
                if head_1 == tail_1:
                    w1 = next_words.pop(0)
                    responses[i]   = f"{responses[i]} {w1}"
                    responses[i+1] = " ".join(next_words)

    return responses

def remove_audio_echoes(text):
    """
    Removes invigilator tape bleed without breaking genuine L2 stutters.
    Splits by any sentence boundary (..., ., ?, !).
    If a 3+ word phrase is repeated, it deletes the first one (the echo).
    """
    import re

    # Split text by ..., ., ?, or ! but keep the punctuation in the list
    parts = re.split(r'(\.\.\.|\.|\?|!)', text)

    # Group the text chunks with their matching punctuation
    chunks = []
    for i in range(0, len(parts) - 1, 2):
        text_part = parts[i].strip()
        punct = parts[i+1]
        if text_part:
            chunks.append((text_part, punct))

    # Handle the very last piece of text if there's no punctuation at the end
    if len(parts) % 2 != 0 and parts[-1].strip():
        chunks.append((parts[-1].strip(), ""))

    if len(chunks) < 2:
        return text # Nothing to compare

    cleaned_chunks = [chunks[0]]

    for i in range(1, len(chunks)):
        prev_text, prev_punct = cleaned_chunks[-1]
        curr_text, curr_punct = chunks[i]

        # ── THE STUTTER THRESHOLD ──
        # Only compare if both phrases are 3+ words long
        if len(prev_text.split()) >= 3 and len(curr_text.split()) >= 3:

            # If they are practically identical (score > 80)
            if fuzzy_score(prev_text, curr_text) > 80:
                # Replace the invigilator's prompt with the student's attempt,
                # keeping the student's ending punctuation.
                cleaned_chunks[-1] = (curr_text, curr_punct)
                continue

        cleaned_chunks.append((curr_text, curr_punct))

    # Recombine the cleaned text and punctuation
    return " ".join([f"{t}{p}" for t, p in cleaned_chunks]).strip()

def remove_cross_sentence_echoes(responses):
    """
    Removes full-phrase echoes that bleed across sentence boundaries.
    Example:
      Sentence 5: "¿Qué dices de qué pasa, héroe?"
      Sentence 6: "¿Qué dices de qué pasa, héroe? Duro seca lezar también."
                  -> becomes "Duro seca lezar también."
    """
    for i in range(1, 30):
        if i in responses and (i + 1) in responses:
            curr_resp = responses[i]
            next_resp = responses[i+1]

            curr_words = curr_resp.split()
            next_words = next_resp.split()

            # 1. Forward phrase bleed (Next sentence STARTS with current sentence)
            if len(curr_words) >= 3 and len(next_words) > len(curr_words):
                head_of_next = " ".join(next_words[:len(curr_words)])
                # If they are practically identical, chop the echo off the next sentence
                if fuzzy_score(curr_resp, head_of_next) > 85:
                    responses[i+1] = " ".join(next_words[len(curr_words):]).strip()
                    # Update variables in case we need them for the backward check
                    next_resp = responses[i+1]
                    next_words = next_resp.split()

            # 2. Backward phrase bleed (Current sentence ENDS with next sentence)
            if len(next_words) >= 3 and len(curr_words) > len(next_words):
                tail_of_curr = " ".join(curr_words[-len(next_words):])
                # If they are practically identical, chop the echo off the current sentence
                if fuzzy_score(next_resp, tail_of_curr) > 85:
                    responses[i] = " ".join(curr_words[:-len(next_words)]).strip()

    return responses

def extract_responses(segments, spanish_stimuli):
    responses = {}
    confident_anchor = 0  # The highest sentence index we've mapped CONFIDENTLY

    print_header("  DYNAMIC ALIGNMENT (ANCHOR-BASED)")

    for i, seg in enumerate(segments):
        if should_skip(seg["text"]):
            print(f"  {color('SKIP', 'dim')}  #{i:03d}  [filtered]  {seg['text'][:60]}")
            continue

        parts = split_merged_segment(seg["text"], spanish_stimuli)

        for part in parts:
            best_idx  = -1
            best_score = -100

            # ── 1. SCORE EVERY SENTENCE ──
            for j, stim in enumerate(spanish_stimuli):
                raw_score = fuzzy_score(part, stim)
                bonus = 0

                # Bonus for chronological progression from our Anchor
                if j == confident_anchor:     bonus = 10
                elif j == confident_anchor+1: bonus = 20  # Strongly favor the next logical sentence
                elif j == confident_anchor+2: bonus = 10

                # Distance-based penalty to prevent ghost echoes.
                # If anchor is 10, guessing sentence 9 gets -4 penalty. Guessing sentence 2 gets -32.
                if j < confident_anchor:
                    bonus -= (confident_anchor - j) * 4

                final_score = raw_score + bonus
                if final_score > best_score:
                    best_score = final_score
                    best_idx   = j

            actual_fuzzy = fuzzy_score(part, spanish_stimuli[best_idx])

            # ── 2. MOVE THE ANCHOR ONLY IF WE ARE SURE ──
            # This prevents a bad guess from destroying the rest of the file's alignment.
            if actual_fuzzy >= 55 and best_idx > confident_anchor:
                # Prevent runaway jumps (max 3 spots at a time) to contain chaos
                if best_idx <= confident_anchor + 3:
                    confident_anchor = best_idx

            target_idx = best_idx
            sentence_num = target_idx + 1
            stimulus = spanish_stimuli[target_idx]

            if target_idx >= 30:
                print(f"  {color('DROP (END)', 'dim')}  #{i:03d}  [past sentence 30] {part[:40]}...")
                continue

            # ── 3. APPLY TO RESPONSES ──
            if sentence_num in responses:
                old_text     = responses[sentence_num]
                score_old    = fuzzy_score(old_text, stimulus)
                score_new    = actual_fuzzy
                merged       = old_text + " " + part
                score_merged = fuzzy_score(merged, stimulus)

                is_time_travel = target_idx < confident_anchor - 1

                if is_time_travel:
                    # Lock down past sentences to prevent stutters from ruining perfect matches
                    if score_new > score_old + 20:
                        responses[sentence_num] = part
                        print(f"  {color('OVERWRITE (PAST)', 'green')}  #{i:03d}  →  sentence {sentence_num:02d} (massive improvement)")
                    else:
                        print(f"  {color('IGNORE (PAST ECHO)', 'dim')}  #{i:03d}  [ignored for {sentence_num:02d}] {part[:40]}...")
                else:
                    # Normal updating for active sentences
                    if score_merged >= max(score_old, score_new) - 5:
                        responses[sentence_num] = merged
                        print(f"  {color('CONCAT', 'yellow')}  #{i:03d}  →  sentence {sentence_num:02d}")
                    elif score_new > score_old + 5:
                        responses[sentence_num] = part
                        print(f"  {color('OVERWRITE', 'green')}  #{i:03d}  →  sentence {sentence_num:02d} (improved match)")
                    else:
                        print(f"  {color('IGNORE (WEAK)', 'dim')}  #{i:03d}  [ignored for {sentence_num:02d}] {part[:40]}...")
            else:
                responses[sentence_num] = part
                tag = 'ASSIGN' if actual_fuzzy >= 45 else 'GUESS'
                col = 'blue'   if actual_fuzzy >= 45 else 'dim'
                print(f"  {color(tag, col)}  #{i:03d}  →  sentence {sentence_num:02d} (score: {actual_fuzzy:.1f})")

    # Cleanups
    responses = fix_bleed_over(responses, spanish_stimuli)

    # Check if remove_cross_sentence_echoes exists before calling it (from the previous step)
    if 'remove_cross_sentence_echoes' in globals():
        responses = remove_cross_sentence_echoes(responses)

    for key in responses:
        responses[key] = remove_audio_echoes(responses[key])

    return responsesdef write_to_excel(excel_path, file_id, responses):
    """
    Write responses into column C of the matching participant sheet.
    Any sentence number with no detected response gets '[no response]'.
    """
    import openpyxl
    wb = openpyxl.load_workbook(excel_path)

    # Find the sheet for this participant
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if str(file_id) in sheet_name:
            target_sheet = wb[sheet_name]
            break

    if target_sheet is None:
        print(color(f"    ⚠️  No sheet found containing ID '{file_id}'. Skipping write.", "yellow"))
        return 0

    filled   = 0
    no_resp  = 0

    for row in target_sheet.iter_rows(min_row=2):
        cell_a = row[0].value   # Column A: sentence number

        try:
            sentence_num = int(cell_a)
        except (ValueError, TypeError):
            continue

        if sentence_num < 1 or sentence_num > 30:
            continue

        if sentence_num in responses:
            # ── Response detected — write it ──────────────────────────────
            row[2].value = responses[sentence_num]
            filled += 1
        else:
            # ── No response detected — write [no response] ────────────────
            row[2].value = "[no response]"
            no_resp += 1

    wb.save(excel_path)
    print(f"    ✅ Wrote {filled}/30 responses  +  {no_resp} × '[no response]'  →  sheet '{target_sheet.title}'")
    return filled + no_resp


def process_transcript(transcript_path, file_id, excel_path):
    print(color(f"\n  📄 Parsing: {transcript_path}", "dim"))
    segments = parse_transcript(transcript_path)
    print(f"    ✅ {len(segments)} segments loaded")

    if not segments:
        print(color("    ⚠️  No segments found — skipping", "yellow"))
        return {}

    boundary = find_spanish_boundary(segments)
    spanish  = segments[boundary:]
    print(f"    ✅ {len(spanish)} segments after boundary")

    responses = extract_responses(spanish, EIT_STIMULI_SPANISH)

    print_header(f"  RESPONSES EXTRACTED  ({len(responses)}/30)")
    for n in range(1, 31):
        resp = responses.get(n, color("— not found → [no response]", "yellow"))
        num  = color(f"  {n:02d}.", "dim")
        stim = color(f"  [{EIT_STIMULI_SPANISH[n-1][:45]}]", "dim")
        print(f"{num}{stim}")
        print(f"      → {resp}")

    if os.path.isfile(excel_path):
        print_header("  WRITING TO EXCEL")
        write_to_excel(excel_path, file_id, responses)
    else:
        print(color(f"\n  ⚠️  Excel not found at {excel_path} — skipping write", "yellow"))

    return responses


def infer_participant_id(folder_name):
    m = re.match(r"(\d+)", os.path.basename(folder_name))
    return str(int(m.group(1))) if m else folder_namedef main():
    check_requirements()

    transcripts_root = "./output"
    excel_path       = "./AutoEIT Sample Audio for Transcribing.xlsx"

    if not os.path.isdir(transcripts_root):
        print(color(f"\n❌ Transcripts folder not found: {transcripts_root}\n", "bold"))
        sys.exit(1)

    transcript_files = []
    for folder in sorted(os.listdir(transcripts_root)):
        folder_path = os.path.join(transcripts_root, folder)
        tx_path     = os.path.join(folder_path, "transcript.txt")
        if os.path.isdir(folder_path) and os.path.isfile(tx_path):
            transcript_files.append((folder, tx_path))

    if not transcript_files:
        print(color(f"\n⚠️  No transcript.txt files found in: {transcripts_root}\n", "yellow"))
        sys.exit(0)

    print_banner("AutoEIT TRANSCRIBER")
    print(f"\n  📁 Transcripts : {transcripts_root}/")
    print(f"  📊 Excel       : {excel_path}")
    print(f"  📄 Found       : {len(transcript_files)} transcript(s)\n")

    summary = []
    for folder_name, tx_path in transcript_files:
        participant_id = infer_participant_id(folder_name)
        print_banner(f"{folder_name}  (ID: {participant_id})")
        try:
            responses = process_transcript(tx_path, participant_id, excel_path)
            summary.append((folder_name, "✅", len(responses)))
        except Exception as e:
            print(color(f"\n  ❌ Failed: {e}", "red", "bold"))
            summary.append((folder_name, "❌", 0))

    print_banner("COMPLETE 🎉")
    print(f"\n  {'FILE':<35}  {'STATUS':<6}  RESPONSES")
    print(f"  {'─'*35}  {'─'*6}  {'─'*10}")
    for fname, status, n in summary:
        print(f"  {fname:<35}  {status:<6}  {n}/30")

    passed = sum(1 for _, s, _ in summary if s == "✅")
    print(f"\n  {passed}/{len(summary)} files processed successfully.\n")


if __name__ == "__main__":
    main()